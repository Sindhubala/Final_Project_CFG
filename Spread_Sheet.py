import csv
from openpyxl import load_workbook,Workbook
import numpy as np
import os

def read_data():
    data = []
    with open('sales.csv', 'r') as sales_csv:
        spreadsheet = csv.DictReader(sales_csv)
        for row in spreadsheet:
            print(dict(row))
            data.append(dict(row))
    return data

#Monthly changes as a percentage
def monthly_change(Data):
    monthly_changes=[]
    monthly_changes.append(["Month", "change as percentage"])

    previous_value = 0
    for row in Data:
        sale = int(row["sales"])
        if(previous_value == 0):
            mc_percentage = float((sale / sale) * 100)

        else:
            mc_percentage = float((sale / previous_value) * 100)

        mc_percentage = round(mc_percentage, 2)
        previous_value = sale

        mc_per = ("{} Month change as percentage {}%".format(row["month"],mc_percentage))
        monthly_changes.append([row["month"],mc_percentage])
        print(mc_per)
    write_to_spreadsheet(monthly_changes)

def sale_summary(data):
    #total sale
    sales = []
    summary =[]
    for row in data:
        sale = int(row['sales'])
        sales.append(sale)
    total = sum(sales)

    # The average
    avg = total / 12
    avg = round(avg, 2)
    summary.append(["",""])
    summary.append(["Total Sales",total])
    summary.append(["Average",avg])
    print('Total sales: {}'.format(total))
    print("Average of the year : {}".format(avg))

    # Months with the highest and lowest sales
    maximum = max(sales)
    res = next((sub for sub in data if sub['sales'] == str(maximum)), None)
    print("Highest sale in the year : {} {}".format(res["month"], maximum))

    minimum = min(sales)
    res = next((sub for sub in data if sub['sales'] == str(minimum)), None)
    print("Lower sale in the year : {} {}".format(res["month"], minimum))

    summary.append(["Maximum",maximum])
    summary.append(["Minimum",minimum])
    write_to_spreadsheet(summary)

def write_to_spreadsheet(data):
    ini_array = np.array(data)
    result= tuple(map(tuple,ini_array))
    if(os.path.exists("spreadsheet.xlsx")):
        workbook = load_workbook("spreadsheet.xlsx")
        worksheet = workbook.active
        for row in result:
            worksheet.append(row)
        workbook.save("spreadsheet.xlsx")
    else:
        workbook = Workbook()
        workbook.save("spreadsheet.xlsx")
        worksheet = workbook.active
        for row in result:
            worksheet.append(row)
        workbook.save("spreadsheet.xlsx")



def run():
    data = read_data()
    monthly_change(data)
    sale_summary(data)

run()


